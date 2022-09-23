import eel
import openpyxl
from os.path import exists
from datetime import datetime
import base64
import time
import json
from PIL import Image

# read_path = "excel/test-1.xlsx"
read_path = "excel/read.xlsx"
write_path = "excel/write.xlsx"

dataset = []
test = []
def getData():

    wb = openpyxl.load_workbook(read_path)
    ws = wb.active
    # sheet_names = wb.sheetnames
    rows_cnt = ws.max_row
    cols_cnt = ws.max_column
    # print(rows_cnt)

    start_row = 0 #count real data starts
    flag = False
    for r in range(1, rows_cnt):
        if flag == True:
            break
        for c in range(1, cols_cnt):
            if ws.cell(row=r, column=c).value == "Q.No":
                start_row = r + 1
                flag = True
                break

    flag = False
    for r in range(start_row, rows_cnt+1):
        subset = {}
        ans_set = {}
        if (ws.cell(row=r, column=5).value) is None:
            break
        if ws.cell(row=r, column=1).value:
            flag = False
            subset['order'] = ws.cell(row=r, column=1).value

            subset['question'] = ws.cell(row=r, column=3).value
            subset['type'] = ws.cell(row=r, column=4).value
            question_media = ws.cell(row=r, column=2).value
            
            if question_media != None:
                if ws.cell(row=r, column=4).value == "Q. picture multiple response" or ws.cell(row=r, column=4).value == "Q. picture single response":
                    question_media = getImageData(question_media)
                subset['question_media'] = question_media

            answer = ws.cell(row=r, column=6).value
            if ws.cell(row=r, column=4).value == "Single picture response" or ws.cell(row=r, column=4).value == "Multiple picture response":
                img_path = ws.cell(row=r, column=6).value.split('>')[0]
                img_name = ws.cell(row=r, column=6).value.split('>')[1]
                flag = True
                answer = getImageData(img_path)
                next = ws.cell(row=r, column=7).value
                # ans_set[next] = answer + ">" + img_name
                ans_set[answer + ">" + img_name] = next
                subset['ans_set'] = ans_set
            else:
                next = ws.cell(row=r, column=7).value
                ans_set[answer] = next
                subset['ans_set'] = ans_set
            dataset.append(subset)
        else:
            subset = dataset[len(dataset)-1]
            if flag == True:
                img_path = ws.cell(row=r, column=6).value.split('>')[0]
                img_name = ws.cell(row=r, column=6).value.split('>')[1]
                answer = getImageData(img_path)
                subset['ans_set'][answer + ">" + img_name] = ws.cell(row=r, column=7).value

            else:
                subset['ans_set'][ws.cell(row=r, column=6).value] = ws.cell(row=r, column=7).value

    
    return dataset
    print(json.dumps(dataset, default=lambda o: None))


def getImageData(img_path):
    basewidth = 1366
    image = Image.open("assets/images/"+img_path)
    wpercent = (basewidth / float(image.size[0]))
    hsize = int((float(image.size[1]) * float(wpercent)))
    image = image.resize((basewidth, hsize), Image.ANTIALIAS)
    image.save("assets/images/" + img_path)

    f=open("assets/images/"+img_path, "rb")
    data = base64.b64encode(f.read()).decode('utf-8')

    return data

# getData()

@eel.expose
def getFirstQuestion():
    dataset = getData()
    return dataset[0]

@eel.expose
def getQuestion(index):
    # dataset = getData()
    return dataset[index-1]

@eel.expose 
def writeExcel(write_data_set):
    file_exists = exists(write_path)

    questions = write_data_set[0]
    q_len = len(questions)

    answers = write_data_set[1]
    a_len = len(answers)
    
    qindex = write_data_set[3]
    aindex = write_data_set[4]

    start_time = write_data_set[2]
    now = datetime.now()
    last_time = now.strftime("%d/%m/%Y %H:%M:%S")
    
    if file_exists == True:
        wb = openpyxl.load_workbook(write_path)
        ws = wb.active

        rows_cnt = ws.max_row
        start_line = rows_cnt + 2

        #time setting part
        ws.cell(row=start_line+2, column=2).value = "user" + str(start_line+2)
        ws.cell(row=start_line+2, column=3).value = start_time
        ws.cell(row=start_line+qindex[len(qindex)-1] + 1, column=3).value = last_time

        for r in range(0, q_len):
            ws.cell(row=start_line+qindex[r]+1, column=4).value = questions[r]
            ws.cell(row=start_line+qindex[r]+1, column=1).value = r + 1

        for r in range(0, a_len):
            ws.cell(row=start_line+aindex[r]+1, column=5).value = answers[r]    

        wb.save("excel/write.xlsx")
            

    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        c1= ws.cell(row=1, column=1)
        c1.value = "Sl No" 
        c1= ws.cell(row=1, column=2)
        c1.value = "User name" 
        c1= ws.cell(row=1, column=3)
        c1.value = "Date & time" 
        c1= ws.cell(row=1, column=4)
        c1.value = "Question" 
        c1= ws.cell(row=1, column=5)
        c1.value = "Answer" 

        
        #time setting part
        ws.cell(row=2, column=2).value = "user1"
        ws.cell(row=2, column=3).value = start_time
        ws.cell(row=qindex[len(qindex)-1] + 1, column=3).value = last_time

        for r in range(0, q_len):
            ws.cell(row=qindex[r]+1, column=4).value = questions[r]
            ws.cell(row=qindex[r]+1, column=1).value = r + 1

        for r in range(0, a_len):
            ws.cell(row=aindex[r]+1, column=5).value = answers[r]    

        wb.save("excel/write.xlsx")

@eel.expose
def getDatabase(filename):
    database_path = "excel/" + filename

    wb = openpyxl.load_workbook(database_path)
    ws = wb.active
    # sheet_names = wb.sheetnames
    rows_cnt = ws.max_row
    
    result = []
    for r in range(2, rows_cnt + 1):
        item = ws.cell(row=r, column=1).value
        result.append(item)
    
    return result
    # print(result) 

# getDatabase("Us cities.xlsx")



eel.init('web')

eel.start('index.html', size=(960, 720), port=0)